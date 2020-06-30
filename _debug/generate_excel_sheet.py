import argparse
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font
from faker import Faker

parser = argparse.ArgumentParser()
parser.add_argument("row_count", help="The number of rows to generate")
parser.add_argument("-p", "--with-pseudonym", action="store_true",
            help="Generate a pseudonym")
args = parser.parse_args()
row_count = int(args.row_count)

wb = Workbook()
ws = wb.active
fake = Faker()

col_headers = (
    'RowID',                # A
    'PatientID',            # B
    'PatientName',          # C
    'PatientBirthDate',     # D
    'StudyDate',            # E
    'Modality',             # F
    'Pseudonym'             # G
)

for index, header in enumerate(col_headers):
    ws.cell(row=1, column=index+1).value = header
    ws.cell(row=1, column=index+1).font = Font(bold=True)

for i in range(row_count):
    row = i + 2

    row_id = i + 1
    ws.cell(row=row, column=1).value = row_id

    patient_id = fake.numerify(text='##########')
    ws.cell(row=row, column=2).value = patient_id
    patient_name = f'{fake.last_name()}, {fake.first_name()}'
    ws.cell(row=row, column=3).value = patient_name

    patient_birth_date = fake.date_of_birth(minimum_age=15)
    ws.cell(row=row, column=4).number_format = 'DD.MM.YYYY'
    ws.cell(row=row, column=4).value = patient_birth_date

    study_date = fake.date_between(start_date='-2y', end_date='today')
    ws.cell(row=row, column=5).number_format = 'DD.MM.YYYY'
    ws.cell(row=row, column=5).value = study_date

    modality = fake.random_element(elements=('CT', 'MR', 'DX'))
    ws.cell(row=row, column=6).value = modality

    if args.with_pseudonym:
        pseudonym = fake.lexify(text='????????',
                letters='0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ')
        ws.cell(row=row, column=7).value = pseudonym

ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 16
ws.column_dimensions['C'].width = 32
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 16
ws.column_dimensions['F'].width = 16
ws.column_dimensions['G'].width = 16

parent_folder = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
samples_folder = os.path.join(parent_folder, '_samples')
filename = f'sample_sheet_generated_{row_count}.xlsx'
sample_sheet_path = os.path.join(samples_folder, filename)

wb.save(sample_sheet_path)