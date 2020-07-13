import re
from openpyxl import load_workbook
import datetime

class ExcelError(Exception):
    def __init__(self, message, errors):
        super().__init__(message)
        self.message = message
        self.errors = errors

    def __str__(self):
        return f'{self.message}: {", ".join(self.errors)}'


class ExcelProcessor:
    def __init__(self, excel_file, worksheet=None):
        self._worksheet = worksheet
        self._excel_file = excel_file
        self.wb = None
        self.ws = None
        self._cols = []

        self._open()

    # Find a specific column in an Excel sheet (by a title in the header row)
    def _search_column(self, column_title):
        col = 1
        while True:
            value = self.ws.cell(column=col, row=1).value
            if not value:
                break
            elif value == column_title:
                return col

            col += 1

        return None

    def _scan_columns(self):
        cols = dict()
        cols['RequestID'] = self._search_column('RequestID')
        cols['PatientID'] = self._search_column('PatientID')
        cols['PatientName'] = self._search_column('PatientName')
        cols['PatientBirthDate'] = self._search_column('PatientBirthDate')
        cols['StudyDate'] = self._search_column('StudyDate')
        cols['Modality'] = self._search_column('Modality')
        cols['Pseudonym'] = self._search_column('Pseudonym')
        cols['Exclude'] = self._search_column('Exclude')

        self._check_columns(cols)

        self._cols = cols

    def _check_columns(self, cols):
        required_columns = [
            'RequestID',
            'PatientID',
            'PatientName',
            'PatientBirthDate',
            'StudyDate', 
            'Modality',
            'Pseudonym'
        ]

        for column_name in required_columns:
            if not cols[column_name]:
                self._errors.append(f'{column_name} column missing')

    def _open(self):
        self._errors = []

        self.wb = load_workbook(filename=self._excel_file)

        if self._worksheet is not None:
            self.ws = self.wb[self._worksheet]
        else:
            self.ws = self.wb.active

        self._scan_columns()

        if len(self._errors) > 0:
            raise ExcelError(f'Invalid format of Excel file', self._errors)

    def _scan_rows(self):
        cols = self._cols

        r = 2 # Skip the header row
        while(True):
            request_id = self.ws.cell(column=cols['RequestID'], row=r).value
            patient_id = self.ws.cell(column=cols['PatientID'], row=r).value
            patient_name = self.ws.cell(column=cols['PatientName'], row=r).value
            patient_birth_date = self.ws.cell(column=cols['PatientBirthDate'], row=r).value
            modality = self.ws.cell(column=cols['Modality'], row=r).value
            study_date = self.ws.cell(column=cols['StudyDate'], row=r).value
            pseudonym = self.ws.cell(column=cols['Pseudonym'], row=r).value

            # Check for an empty row and then stop parsing
            if not (request_id or patient_id or patient_name or patient_birth_date
                    or modality or study_date or pseudonym):
                break

            exclude = False
            if cols['Exclude']:
                exclude = bool(self.ws.cell(column=cols['Exclude'], row=r).value)

            if not exclude:
                row = dict()
                row['Row'] = r
                row['RequestID'] = self._clean_request_id(request_id, r)
                row['PatientID'] = self._clean_patient_id(patient_id)
                row['PatientName'] = self._clean_patient_name(patient_name)
                row['PatientBirthDate'] = self._clean_patient_birth_date(patient_birth_date, r)
                row['Modality'] = self._clean_modality(modality, r)
                row['StudyDate'] = self._clean_study_date(study_date, r)
                row['Pseudonym'] = self._clean_pseudonym(pseudonym)
                self._clean_row(row, r)

                self._data.append(row)

            r += 1

    def _clean_request_id(self, request_id, r):
        request_id = str(request_id).strip()
        if request_id and request_id not in self._request_ids:
            self._request_ids.add(request_id)
            return request_id
        elif request_id and request_id in self._request_ids:
            self._errors.append('Duplicate RequestID in row ' % r)  
        else:
            self._errors.append('Missing RequestID in Excel row ' % r)

    def _clean_patient_id(self, patient_id):
        patient_id = str(patient_id).strip()
        return str(patient_id) if patient_id else ''

    def _clean_patient_name(self, patient_name):
        return re.sub(r',\s*', '^', str(patient_name).strip()) if patient_name else ''

    def _clean_patient_birth_date(self, patient_birth_date, r):
        if patient_birth_date and isinstance(patient_birth_date, datetime.datetime):
            return patient_birth_date
        elif patient_birth_date:
            self._errors.append('Invalid PatientBirthDate (row %d)' % r)
        else:
            self._errors.append('Missing PatientBirthDate (row %d)' % r)

    def _clean_modality(self, modality, r):
        if modality:
            return str(modality)
        else:
            self._errors.append('Missing Modality (row %d)' % r)

    def _clean_study_date(self, study_date, r):
        if study_date and isinstance(study_date, datetime.datetime):
            return study_date
        elif study_date:
            self._errors.append('No valid StudyDate (row %d)' % r)
        else:
            self._errors.append('Missing StudyDate (row %d)' % r)

    def _clean_pseudonym(self, pseudonym):
        return str(pseudonym) if pseudonym else ''

    def _clean_row(self, row, r):
        if not(row['PatientID'] 
                or (row['PatientName'] and row['PatientBirthDate'])):
            self._errors.append('Missing PatientID or PatientName/PatientBirthDate (row %d)' % r)

        # TODO check that combination patient_id and pseudonym is unique
        # TODO check that combination patient_name and patient_birth_date and pseudonym is unique
        # TODO check that combination patient_id and patient_name and patient_birth_date is unique

    def extract_data(self):
        self._request_ids = set()
        self._data = []
        self._errors = []

        self._scan_rows()

        if len(self._errors) > 0:
            raise ExcelError(f'Invalid format of Excel file', self._errors)

        return self._data

    def close(self):
        self.wb.close()