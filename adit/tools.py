import os
import string
import random
import re
import pydicom
from openpyxl import load_workbook

class Anonymizer:
    def __init__(
            self,
            anonymize_patient_id=False,
            anonymize_patient_name=True,
            normalize_birth_date=True,
            anonymize_all_person_names=True,
            anomymize_curves=True,
            remove_private_tags=True,
            remove_other_patient_ids=True):

        self.anonymize_patient_id = anonymize_patient_id
        self.anonymize_patient_name = anonymize_patient_name
        self.normalize_birth_date = normalize_birth_date
        self.anonymize_all_person_names = anonymize_all_person_names
        self.anomymize_curves = anomymize_curves
        self.remove_private_tags = remove_private_tags
        self.remove_other_patient_ids = remove_other_patient_ids

        self.generated_pseudonyms = []

    def _person_names_callback(self, ds, data_element):
        if data_element.VR == "PN":
            data_element.value = "Anonymized"

    def _curves_callback(self, ds, data_element):
        if data_element.tag.group & 0xFF00 == 0x5000:
            del ds[data_element.tag]

    def anonymize_dataset(
            self,
            ds,
            patient_id=None,
            patient_name=None):

        if self.anonymize_patient_id:
            ds.PatientID = "Anonymized"

        if self.anonymize_patient_name:
            ds.PatientName = "Anonymized"

        if self.normalize_birth_date:
            birth_date = ds.PatientBirthDate
            ds.PatientBirthDate = birth_date[0:4] + "0101"

        if self.anonymize_all_person_names:
            ds.walk(self._person_names_callback)

        if self.anomymize_curves:
            ds.walk(self._curves_callback)

        if self.remove_private_tags:
            ds.remove_private_tags()

        if self.remove_other_patient_ids:
            if hasattr(ds, 'OtherPatientIDs'):
                del ds.OtherPatientIDs
            if hasattr(ds, 'OtherPatientIDsSequence'):
                del ds.OtherPatientIDsSequence

        if patient_id is not None:
            ds.PatientID = patient_id

        if patient_name is not None:
            ds.PatientName = patient_name

        return (ds.PatientID, ds.PatientName)

    def anonymize_folder(self, folder, *args, **kwargs):
        if not os.path.isdir(folder):
            raise ValueError(f"Invalid folder: {folder}")

        callback = kwargs.pop('callback', None)
        if callback is not None and not callable(callback):
            raise ValueError("callback must be callable function")

        for root, _, files in os.walk(folder):
            for filename in files:
                filepath = os.path.join(root, filename)
                ds = pydicom.dcmread(filepath)
                self.anonymize_dataset(ds, *args, **kwargs)

                # Allow to manipulate the dataset before saving it to disk
                if callback:
                    callback(ds)

                ds.save_as(filepath)

        return True

    def generate_pseudonym(
        self,
        length=12,
        prefix="",
        suffix="",
        random_string=True,
        uppercase=True
    ):
        pseudonym = None
        while True:
            chars = string.ascii_lowercase + string.digits
            if random_string:
                pseudonym = ''.join(random.choice(chars) for _ in range(length))
                if uppercase:
                    pseudonym = pseudonym.upper()
            else:
                pseudonym = str(len(self.generated_pseudonyms) + 1).zfill(length)
            pseudonym = prefix + pseudonym + suffix
            if not pseudonym in self.generated_pseudonyms:
                self.generated_pseudonyms.append(pseudonym)
                return pseudonym


class ExcelError(Exception):
    pass


class ExcelProcessor:
    PATIENT_ID_COL = "patient_id_col"
    PATIENT_NAME_COL = "patient_name_col"
    PATIENT_BIRTH_DATE_COL = "patient_birth_date_col"
    STUDY_DATE_COL = "study_date_col"
    MODALITY_COL = "modality_col"
    PSEUDONYM_COL = "pseudonym_col"
    EXCLUDE_COL = "exclude_col"
    STATUS_COL = "status_col"

    def __init__(self, filename):
        self.filename = filename
        self.wb = None
        self.ws = None
        self.cols = []
        self.data = []

    # Find a specific column in an Excel sheet (by a title in the header row)
    def _search_column(self, column_title):
        col = 1
        while True:
            value = self.ws.cell(column=col, row=1).value
            if not value:
                break
            elif value == column_title:
                return col

            col += 1

        return None

    def _scan_columns(self):
        cols = dict()
        cols[self.PATIENT_ID_COL] = self._search_column("PatientID")
        cols[self.PATIENT_NAME_COL] = self._search_column("PatientName")
        cols[self.PATIENT_BIRTH_DATE_COL] = self._search_column("PatientBirthDate")
        cols[self.STUDY_DATE_COL] = self._search_column("StudyDate")
        cols[self.MODALITY_COL] = self._search_column("Modality")
        cols[self.PSEUDONYM_COL] = self._search_column("Pseudonym")
        cols[self.EXCLUDE_COL] = self._search_column("Exclude")
        cols[self.STATUS_COL] = self._search_column("Status")

        self._check_columns(cols)

        self.cols = cols

    def _check_columns(self, cols):
        required_columns = [
            self.PATIENT_ID_COL,
            self.PATIENT_NAME_COL,
            self.PATIENT_BIRTH_DATE_COL,
            self.STUDY_DATE_COL, 
            self.MODALITY_COL,
            self.PSEUDONYM_COL, 
            self.STATUS_COL
        ]

        for column_name in required_columns:
            if not cols[column_name]:
                raise ExcelError(f"Invalid Excel format: {column_name} column missing")

    def _scan_rows(self):
        cols = self.cols
        data = []
        r = 2 # Skip the header row
        while(True):
            exclude = bool(self.ws.cell(column=cols[self.EXCLUDE_COL], row=r).value)

            if not exclude:
                row = dict()
                row['RowID'] = r
                patient_id = self.ws.cell(column=cols[self.PATIENT_ID_COL], row=r).value
                row['PatientID'] = str(patient_id) if patient_id else ''
                patient_name = self.ws.cell(column=cols[self.PATIENT_NAME_COL], row=r).value
                row['PatientName'] = re.sub(r',\s*', '^', str(patient_name).strip()) if patient_name else ''
                birth_date = self.ws.cell(column=cols[self.PATIENT_BIRTH_DATE_COL], row=r).value
                row['PatientBirthDate'] = birth_date.strftime('%Y%m%d') if birth_date else ''
                modality = self.ws.cell(column=cols[self.MODALITY_COL], row=r).value
                row['Modality'] = str(modality) if modality else ''
                study_date = self.ws.cell(column=cols[self.STUDY_DATE_COL], row=r).value
                row['StudyDate'] = study_date.strftime('%Y%m%d') if study_date else ''
                pseudonym = self.ws.cell(column=cols[self.PSEUDONYM_COL], row=r).value
                row['Pseudonym'] = str(pseudonym) if pseudonym else ''

                row_empty = True
                for key, val in row.items():
                    if key != 'RowID' and val:
                        row_empty = False

                if not row_empty:
                    self._check_row(row)
                    data.append(row)
                else:
                    break

            r += 1

        self.data = data

    def _check_row(self, row):
        row_id = row['RowID']
        if not(row['PatientID'] or (row['PatientName'] and row['PatientBirthDate'])):
            raise ExcelError("Missing PatientID or PatientName and PatientBirthDate on row %d" % row_id)
        if not row['StudyDate']:
            raise ExcelError("Missing StudyDate on row %d" % row_id)
        if not row['Modality']:
            raise ExcelError("Missing Modality on row %d" % row_id)

    def open(self, worksheet_name):
        self.wb = load_workbook(filename=self.filename)
        self.ws = self.wb[worksheet_name]
        self._scan_columns()
        self._scan_rows()

    def close(self):
        self.wb.close()

    def save(self):
        self.wb.save(filename=self.filename)

    def set_cell_value(self, column_name, row, value):
        column = self.cols[column_name]
        self.ws.cell(column=column, row=row, value=value)